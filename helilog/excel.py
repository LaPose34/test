"""Entrada y salida en Excel.

Entrada: un libro con las hojas
  - Config          parámetros generales (clave / valor)
  - Puntos          helipuntos con coordenadas y restricciones
  - Distancias      matriz de km entre puntos (opcional; si falta, haversine)
  - Helicopteros    la flota
  - Requerimientos  demanda por grupos (PAX o CARGA), divisible en viajes
  - Pasajeros       personas individuales con peso + equipaje (opcional)

Salida: un libro con la programación de vuelo por rotaciones (al estilo del
software anterior), un resumen de la flota y el eco de los requerimientos.

Requiere `openpyxl` (solo para las funciones de este módulo).
"""

from __future__ import annotations

from datetime import datetime, timedelta

from . import catalog
from .models import Helicopter, Helipad, Passenger, Scenario, TransportRequest
from .optimizer import Leg, RoutePlan


def _yes(value, default=False) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().upper() in ("SI", "SÍ", "S", "YES", "Y", "TRUE", "1", "X")


def _num(value, default=None):
    if value is None or str(value).strip() == "":
        return default
    return float(value)


def _rows(ws, min_row):
    """Filas no vacías como listas de valores."""
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            yield list(row)


# --------------------------------------------------------------------- lectura

CONFIG_KEYS = {
    "peso estandar pax": "pax_weight_kg",
    "peso estándar pax": "pax_weight_kg",
    "precio combustible": "fuel_price_per_l",
    "densidad combustible": "fuel_density_kg_per_l",
    "altitud de crucero": "cruise_altitude_m",
    "regreso a base": "return_to_base",
    "hora inicio": "start_time",
}


def scenario_from_excel(path: str) -> tuple[Scenario, str]:
    """Lee un escenario desde Excel. Devuelve (escenario, hora_inicio)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {name.strip().lower(): wb[name] for name in wb.sheetnames}

    def sheet(*names):
        for n in names:
            if n in sheets:
                return sheets[n]
        return None

    # Config
    params = {"start_time": "09:00"}
    ws = sheet("config")
    if ws is not None:
        for row in _rows(ws, 1):
            key = str(row[0] or "").strip().lower()
            for prefix, field_name in CONFIG_KEYS.items():
                if key.startswith(prefix):
                    value = row[1]
                    if field_name == "return_to_base":
                        params[field_name] = _yes(value)
                    elif field_name == "start_time":
                        if value is not None:
                            params[field_name] = str(value).strip()[:5]
                    else:
                        params[field_name] = _num(value)
    start_time = params.pop("start_time")

    # Puntos: Id | Nombre | Lat | Lon | Tamaño | Peso máx (kg) | Combustible | Prohibidos
    ws = sheet("puntos", "helipuntos", "helipads")
    if ws is None:
        raise ValueError("Falta la hoja 'Puntos'")
    helipads: dict[str, Helipad] = {}
    for row in _rows(ws, 2):
        pid = str(row[0] or "").strip()
        if not pid or pid.lower() == "id":
            continue
        banned = str(row[7] or "") if len(row) > 7 else ""
        helipads[pid] = Helipad(
            id=pid,
            name=str(row[1] or pid).strip() if len(row) > 1 else pid,
            lat=_num(row[2]) if len(row) > 2 else None,
            lon=_num(row[3]) if len(row) > 3 else None,
            size_class=int(_num(row[4], 3)) if len(row) > 4 else 3,
            max_weight_kg=_num(row[5]) if len(row) > 5 else None,
            has_fuel=_yes(row[6]) if len(row) > 6 else False,
            banned_helicopters=frozenset(
                x.strip() for x in banned.replace(";", ",").split(",") if x.strip()
            ),
        )

    # Distancias: matriz con ids en primera fila y primera columna
    distances: dict[tuple[str, str], float] = {}
    ws = sheet("distancias", "distancias (km)", "km")
    if ws is not None:
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        if grid:
            cols = [str(v).strip() if v is not None else "" for v in grid[0][1:]]
            for row in grid[1:]:
                a = str(row[0] or "").strip()
                if not a:
                    continue
                for j, cell in enumerate(row[1 : len(cols) + 1]):
                    b = cols[j] if j < len(cols) else ""
                    if b and a != b and cell is not None and str(cell).strip() != "":
                        distances[(a, b)] = float(cell)

    # Helicopteros: Id | Nombre | Base | Pax | Peso vacío | MTOW | Cabina máx |
    #               Consumo L/h | Tanque L | Velocidad | Precio/h | Pax+carga | Tamaño
    ws = sheet("helicopteros", "helicópteros", "flota")
    if ws is None:
        raise ValueError("Falta la hoja 'Helicopteros'")
    helicopters: list[Helicopter] = []
    for row in _rows(ws, 2):
        hid = str(row[0] or "").strip()
        if not hid or hid.lower() == "id":
            continue
        name = str(row[1] or hid).strip()
        # Si el nombre coincide con un modelo del catálogo, las celdas
        # técnicas vacías se completan con las especificaciones del modelo
        # (capacidad pax permitida en Perú, MTOW, tanque, consumo, etc.).
        specs = catalog.find(name) or {}

        def cell(i, key, default=None):
            v = row[i] if len(row) > i else None
            if v is None or str(v).strip() == "":
                return specs.get(key, default)
            return float(v)

        helicopters.append(
            Helicopter(
                id=hid,
                name=name,
                base=str(row[2] or "").strip(),
                pax_capacity=int(cell(3, "pax_capacity", 0)),
                empty_weight_kg=cell(4, "empty_weight_kg", 0.0),
                mtow_kg=cell(5, "mtow_kg", 0.0),
                max_payload_kg=cell(6, "max_payload_kg"),
                fuel_consumption_lph=cell(7, "fuel_consumption_lph", 0.0),
                fuel_capacity_l=cell(8, "fuel_capacity_l", 0.0),
                cruise_speed_kmh=cell(9, "cruise_speed_kmh", 0.0),
                price_per_hour=cell(10, "price_per_hour", 0.0),
                can_combine_pax_cargo=_yes(row[11], True) if len(row) > 11 else True,
                size_class=int(cell(12, "size_class", 2)),
                vertical_speed_ms=cell(13, "vertical_speed_ms", 0.0),
            )
        )

    # Requerimientos: Req | Origen | Destino | Tipo | Cant | kg/U | Divisible |
    #                 Descripción | Empresa | Proyecto
    requests: list[TransportRequest] = []
    ws = sheet("requerimientos", "requests")
    if ws is not None:
        for row in _rows(ws, 2):
            rid = str(row[0] or "").strip()
            if not rid or rid.lower() in ("req.", "req", "id"):
                continue
            tipo = str(row[3] or "PAX").strip().upper()
            cant = int(_num(row[4], 1))
            kg_u = _num(row[5], 0.0)
            is_pax = tipo in ("PAX", "P")
            requests.append(
                TransportRequest(
                    id=rid,
                    origin=str(row[1] or "").strip(),
                    destination=str(row[2] or "").strip(),
                    pax=cant if is_pax else 0,
                    pax_weight_kg=cant * kg_u if is_pax and kg_u else None,
                    cargo_kg=0.0 if is_pax else cant * kg_u,
                    splittable=_yes(row[6], cant > 1) if len(row) > 6 else cant > 1,
                    units=cant,
                    description=str(row[7] or "").strip() if len(row) > 7 else "",
                    company=str(row[8] or "").strip() if len(row) > 8 else "",
                    project=str(row[9] or "").strip() if len(row) > 9 else "",
                )
            )

    # Pasajeros: Id | Nombre | Peso | Equipaje | Origen | Destino | Escalas
    ws = sheet("pasajeros", "passengers")
    if ws is not None:
        for row in _rows(ws, 2):
            pid = str(row[0] or "").strip()
            if not pid or pid.lower() == "id":
                continue
            via = str(row[6] or "") if len(row) > 6 else ""
            passenger = Passenger(
                id=pid,
                name=str(row[1] or pid).strip(),
                weight_kg=_num(row[2], 0.0),
                baggage_kg=_num(row[3], 0.0),
                origin=str(row[4] or "").strip(),
                destination=str(row[5] or "").strip(),
                via=tuple(x.strip() for x in via.replace(";", ",").split(",") if x.strip()),
            )
            requests.extend(passenger.to_requests())

    scenario = Scenario(
        helipads=helipads,
        helicopters=helicopters,
        requests=requests,
        distances=distances,
        **params,
    )
    scenario.validate()
    return scenario, start_time


# -------------------------------------------------------------------- plantilla

def write_template(path: str) -> None:
    """Genera un libro de entrada con las hojas y encabezados esperados."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Config"
    for r, (k, v) in enumerate(
        [
            ("Peso estándar pax (kg)", 90),
            ("Precio combustible (por L)", 0),
            ("Densidad combustible (kg/L)", 0.8),
            ("Altitud de crucero (m)", 300),
            ("Regreso a base", "SI"),
            ("Hora inicio", "09:00"),
        ],
        1,
    ):
        ws.cell(r, 1, k).font = bold
        ws.cell(r, 2, v)

    def add(name, headers, rows=()):
        w = wb.create_sheet(name)
        for c, h in enumerate(headers, 1):
            w.cell(1, c, h).font = bold
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                w.cell(r, c, v)
        return w

    add(
        "Puntos",
        ["Id", "Nombre", "Lat", "Lon", "Tamaño (1-3)", "Peso máx heli (kg)",
         "Combustible (SI/NO)", "Helis prohibidos"],
        [["MALV", "Base Malvinas", None, None, 3, None, "SI", ""]],
    )
    dist = wb.create_sheet("Distancias")
    dist.cell(1, 1, "km").font = bold
    dist.cell(1, 2, "MALV")
    dist.cell(2, 1, "MALV")
    add(
        "Helicopteros",
        ["Id", "Nombre", "Base", "Pax", "Peso vacío (kg)", "MTOW (kg)",
         "Cabina máx (kg)", "Consumo (L/h)", "Tanque (L)", "Velocidad (km/h)",
         "Precio por hora", "Pax+carga juntos (SI/NO)", "Tamaño (1-3)",
         "Vel. vertical (m/s)"],
        [["OB2106", "Bell 412EP", "MALV", None, None, None, 1000, None, None,
          None, 3500, "SI", None, None]],
    )
    add(
        "Requerimientos",
        ["Req", "Origen", "Destino", "Tipo (PAX/CARGA)", "Cant", "kg/U",
         "Divisible (SI/NO)", "Descripción", "Empresa", "Proyecto"],
    )
    add(
        "Pasajeros",
        ["Id", "Nombre", "Peso (kg)", "Equipaje (kg)", "Origen", "Destino",
         "Escalas (via)"],
    )
    wb.save(path)


# --------------------------------------------------------------------- salida

def _fmt_h(hours: float) -> str:
    m = round(hours * 60)
    return f"{m // 60:02d}H{m % 60:02d}"


def _visits(legs: list[Leg]):
    """Agrupa los tramos en visitas: (pad, [acciones], hora_llegada_horas)."""
    visits = []
    clock = 0.0
    for leg in legs:
        clock += leg.hours
        if leg.km > 0 or not visits:
            visits.append({"pad": leg.to_pad, "legs": [leg], "t": clock,
                           "km": leg.km, "hours": leg.hours, "cost": leg.cost,
                           "from": leg.from_pad})
        else:
            visits[-1]["legs"].append(leg)
    return visits


def write_program(
    path: str,
    scn: Scenario,
    plans: list[RoutePlan],
    requests_by_id: dict[str, TransportRequest] | None = None,
    start_time: str = "09:00",
) -> None:
    """Escribe la programación de vuelo a Excel (estilo rotaciones)."""
    import openpyxl
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb = openpyxl.Workbook()

    # ---- Resumen de flota
    ws = wb.active
    ws.title = "Resumen"
    ws.cell(1, 1, "Optimización logística — ranking por costo").font = bold
    headers = ["Helicóptero", "Viable", "Método", "Costo total", "Horas vuelo",
               "Km", "Motivo si no viable"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h).font = bold
    heli_by_id = {h.id: h for h in scn.helicopters}
    for r, plan in enumerate(plans, 4):
        heli = heli_by_id[plan.helicopter_id]
        ws.cell(r, 1, f"{heli.name or heli.id} ({heli.id})")
        ws.cell(r, 2, "SI" if plan.feasible else "NO")
        ws.cell(r, 3, plan.method if plan.feasible else "")
        if plan.feasible:
            ws.cell(r, 4, round(plan.total_cost, 2))
            ws.cell(r, 5, _fmt_h(plan.total_hours))
            ws.cell(r, 6, round(plan.total_km, 1))
        ws.cell(r, 7, plan.infeasible_reason)

    # ---- Programación del mejor plan viable, por rotaciones
    best = next((p for p in plans if p.feasible), None)
    ws = wb.create_sheet("Programación")
    row = 1
    if best is None:
        ws.cell(1, 1, "Sin plan viable").font = bold
    else:
        heli = heli_by_id[best.helicopter_id]
        base = heli.base
        t0 = datetime.strptime(start_time, "%H:%M")

        def hora(hours_from_start: float) -> str:
            t = t0 + timedelta(hours=hours_from_start)
            return t.strftime("%HH%M")

        visits = _visits(best.legs)
        # separar en rotaciones: cada regreso VOLADO a la base cierra una
        rotations: list[list[dict]] = [[]]
        for v in visits:
            rotations[-1].append(v)
            if v["pad"] == base and v["km"] > 0:
                rotations.append([])
        if rotations and not rotations[-1]:
            rotations.pop()

        req_headers = ["Req.", "E/D", "Ubic.", "Dest/Orig", "Tipo", "#U",
                       "Total kg", "Hora", "Descripción", "Empresa", "Proyecto"]

        for rot in rotations:
            hours_rot = sum(v["hours"] for v in rot)
            cost_rot = sum(v["cost"] for v in rot)
            ws.cell(row, 1, "Helicóptero:").font = bold
            ws.cell(row, 2, f"{heli.name or heli.id} ({heli.id})")
            ws.cell(row + 1, 1, "Tiempo horómetro:").font = bold
            ws.cell(row + 1, 2, _fmt_h(hours_rot))
            ws.cell(row + 2, 1, "Costo:").font = bold
            ws.cell(row + 2, 2, round(cost_rot, 2))
            row += 3
            for c, h in enumerate(req_headers, 1):
                ws.cell(row, c, h).font = bold
            row += 1

            for v in rot:
                if v["km"] > 0:
                    ws.cell(row, 1, f"De {v['from']} a {v['pad']}")
                    ws.cell(
                        row, 2,
                        f"{v['km']:.1f} km — vuelo estimado {_fmt_h(v['hours'])}",
                    )
                    row += 1
                for leg in v["legs"]:
                    action, _, rid = leg.action.partition(" ")
                    if action not in ("recoger", "entregar"):
                        continue
                    req = (requests_by_id or {}).get(rid)
                    base_req = (requests_by_id or {}).get(rid.split("/")[0])
                    info = req or base_req
                    is_e = action == "recoger"
                    if req is not None:
                        n_units = req.units or req.pax
                        total_kg = scn.request_payload_kg(req)
                        tipo = "PAX" if req.pax else "CARGA"
                        other = req.destination if is_e else req.origin
                    else:
                        n_units, total_kg, tipo, other = None, None, "", ""
                    ws.cell(row, 1, rid)
                    ws.cell(row, 2, "E" if is_e else "D")
                    ws.cell(row, 3, v["pad"])
                    ws.cell(row, 4, ("DEST: " if is_e else "ORIG: ") + str(other))
                    ws.cell(row, 5, tipo)
                    ws.cell(row, 6, n_units)
                    ws.cell(row, 7, round(total_kg, 1) if total_kg is not None else None)
                    ws.cell(row, 8, hora(v["t"]))
                    if info is not None:
                        ws.cell(row, 9, info.description)
                        ws.cell(row, 10, info.company)
                        ws.cell(row, 11, info.project)
                    row += 1
            last = rot[-1]
            ws.cell(row, 1, f"Fin de rotación en {last['pad']}")
            ws.cell(row, 2, f"Hora estimada: {hora(last['t'])}")
            row += 2

    # ---- Eco de requerimientos
    ws = wb.create_sheet("Requerimientos")
    headers = ["Req", "Origen", "Destino", "Tipo", "Cant", "Total kg",
               "Divisible", "Descripción", "Empresa", "Proyecto"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = bold
    for r, req in enumerate(scn.requests, 2):
        ws.cell(r, 1, req.id)
        ws.cell(r, 2, req.origin)
        ws.cell(r, 3, req.destination)
        ws.cell(r, 4, "PAX" if req.pax else "CARGA")
        ws.cell(r, 5, req.units or req.pax or None)
        ws.cell(r, 6, round(scn.request_payload_kg(req), 1))
        ws.cell(r, 7, "SI" if req.splittable else "NO")
        ws.cell(r, 8, req.description)
        ws.cell(r, 9, req.company)
        ws.cell(r, 10, req.project)

    wb.save(path)
